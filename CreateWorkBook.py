# ==========================================================
# EQUIPMENT MANAGEMENT SYSTEM
# PART 1
# Imports, Workbook, Theme
# ==========================================================

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Border,
    Side,
    Alignment
)

from openpyxl.worksheet.table import (
    Table,
    TableStyleInfo
)

from openpyxl.worksheet.datavalidation import (
    DataValidation
)

from openpyxl.utils import get_column_letter

from openpyxl.chart import (
    BarChart,
    Reference
)

from datetime import datetime



# ==========================================================
# CONFIGURATION
# ==========================================================

OUTPUT_FILE = "Equipment Management System.xlsx"

MONTH_NAME = "July"

FINANCIAL_SHEET = f"{MONTH_NAME} Financial Telemetry"



# ==========================================================
# CREATE WORKBOOK
# ==========================================================

wb = Workbook()

master = wb.active
master.title = "Master Inventory"

financial = wb.create_sheet(FINANCIAL_SHEET)

dashboard = wb.create_sheet("Dashboard")



# ==========================================================
# CYBER COLOUR PALETTE
# ==========================================================

BACKGROUND = "08121E"

PANEL = "13283E"

HEADER = "00E5FF"

CARD = "0D2538"

GRID = "274A66"

WHITE = "FFFFFF"

GREEN = "00C853"

ORANGE = "FF9800"

RED = "FF3D3D"



# ==========================================================
# COMMON FONTS
# ==========================================================

TITLE_FONT = Font(
    name="Aptos Display",
    size=22,
    bold=True,
    color=HEADER
)

HEADER_FONT = Font(
    name="Aptos",
    size=11,
    bold=True,
    color=WHITE
)

NORMAL_FONT = Font(
    name="Aptos",
    size=10,
    color=WHITE
)

BIG_FONT = Font(
    name="Aptos Display",
    size=26,
    bold=True,
    color=HEADER
)



# ==========================================================
# COMMON BORDER
# ==========================================================

CYAN_SIDE = Side(
    style="thin",
    color=HEADER
)

BORDER = Border(
    left=CYAN_SIDE,
    right=CYAN_SIDE,
    top=CYAN_SIDE,
    bottom=CYAN_SIDE
)



# ==========================================================
# COMMON FILLS
# ==========================================================

BG_FILL = PatternFill(
    "solid",
    fgColor=BACKGROUND
)

HEADER_FILL = PatternFill(
    "solid",
    fgColor=PANEL
)

GROUP_FILL = PatternFill(
    "solid",
    fgColor=CARD
)

# ==========================================================
# JULY FINANCIAL TELEMETRY
# ==========================================================

ws = financial

# ----------------------------------------------------------
# Dark Background
# ----------------------------------------------------------

for col in range(1, 16):
    ws.column_dimensions[get_column_letter(col)].width = 15

for row in range(1, 600):
    for col in range(1, 16):
        ws.cell(row, col).fill = BG_FILL

ws.sheet_view.showGridLines = False

ws.freeze_panes = "A4"


# ----------------------------------------------------------
# Title
# ----------------------------------------------------------

ws.merge_cells("A1:N1")

title = ws["A1"]

title.value = f"{MONTH_NAME.upper()} FINANCIAL TELEMETRY"

title.font = TITLE_FONT

title.alignment = Alignment(
    horizontal="center",
    vertical="center"
)

title.fill = BG_FILL

ws.row_dimensions[1].height = 34


# ----------------------------------------------------------
# Main Headers
# ----------------------------------------------------------

headers = [

("A2:A3","Date"),

("B2:B3","Asset Type"),

("C2:C3","Type of Works"),

("D2:E2","Spare Parts & External Services"),

("F2:G2","Materials"),

("H2:I2","Labour"),

("J2:K2","Transportation"),

("L2:L3","Total"),

("M2:M3","Remarks"),

("N2:N3","Status")

]

for rng,text in headers:

    ws.merge_cells(rng)

    c = ws[rng.split(":")[0]]

    c.value = text

    c.font = HEADER_FONT

    c.fill = HEADER_FILL

    c.border = BORDER

    c.alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )


# ----------------------------------------------------------
# Qty / Cost Sub Headers
# ----------------------------------------------------------

subs = {

"D3":"Qty",
"E3":"Cost",

"F3":"Qty",
"G3":"Cost",

"H3":"Qty",
"I3":"Cost",

"J3":"Qty",
"K3":"Cost"

}

for ref,text in subs.items():

    c = ws[ref]

    c.value = text

    c.font = HEADER_FONT

    c.fill = GROUP_FILL

    c.border = BORDER

    c.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )


# ----------------------------------------------------------
# Column Widths
# ----------------------------------------------------------

widths={

"A":14,

"B":18,

"C":18,

"D":10,

"E":14,

"F":10,

"G":14,

"H":10,

"I":14,

"J":10,

"K":14,

"L":16,

"M":30,

"N":16

}

for col,width in widths.items():

    ws.column_dimensions[col].width=width


START_ROW=4

END_ROW=504

# ==========================================================
# DROPDOWNS
# ==========================================================

asset_validation = DataValidation(
    type="list",
    formula1='"Spider Crane,Boom Lift,Scissor Lift"',
    allow_blank=True
)

works_validation = DataValidation(
    type="list",
    formula1='"Maintenance,Repair"',
    allow_blank=True
)

status_validation = DataValidation(
    type="list",
    formula1='"Pending,Completed"',
    allow_blank=True
)

ws.add_data_validation(asset_validation)
ws.add_data_validation(works_validation)
ws.add_data_validation(status_validation)


# ==========================================================
# FORMAT DATA ROWS
# ==========================================================

for row in range(START_ROW, END_ROW + 1):

    # -----------------------
    # Date
    # -----------------------

    ws[f"A{row}"].number_format = "dd/mm/yyyy"

    # -----------------------
    # Qty
    # -----------------------

    for col in ("D","F","H","J"):
        ws[f"{col}{row}"].number_format = "0"

    # -----------------------
    # Cost
    # -----------------------

    for col in ("E","G","I","K","L"):
        ws[f"{col}{row}"].number_format = '"$"#,##0.00'

    # -----------------------
    # Total Formula
    # -----------------------

    ws[f"L{row}"] = (
        f"=SUM(E{row},G{row},I{row},K{row})"
    )

    # -----------------------
    # Borders / Fonts
    # -----------------------

    for col in range(1,15):

        c = ws.cell(row,col)

        c.font = NORMAL_FONT

        c.border = BORDER

        c.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # -----------------------
    # Alternate Row Colour
    # -----------------------

    if row % 2 == 0:

        fill = PatternFill(
            "solid",
            fgColor="0A2136"
        )

    else:

        fill = PatternFill(
            "solid",
            fgColor="102A43"
        )

    for col in range(1,15):

        ws.cell(row,col).fill = fill

    # -----------------------
    # Drop Downs
    # -----------------------

    asset_validation.add(ws[f"B{row}"])

    works_validation.add(ws[f"C{row}"])

    status_validation.add(ws[f"N{row}"])


# ==========================================================
# TABLE
# ==========================================================

table = Table(
    displayName="FinancialTelemetry",
    ref=f"A3:N{END_ROW}"
)

style = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=False,
    showColumnStripes=False
)

table.tableStyleInfo = style

ws.add_table(table)


# ==========================================================
# AUTO FILTER
# ==========================================================

ws.auto_filter.ref = f"A3:N{END_ROW}"


# ==========================================================
# MONTHLY TOTAL
# ==========================================================

summary = END_ROW + 3

ws[f"K{summary}"] = "Monthly Total"

ws[f"K{summary}"].font = HEADER_FONT

ws[f"K{summary}"].fill = HEADER_FILL

ws[f"L{summary}"] = f"=SUM(L{START_ROW}:L{END_ROW})"

ws[f"L{summary}"].font = HEADER_FONT

ws[f"L{summary}"].fill = HEADER_FILL

ws[f"L{summary}"].number_format = '"$"#,##0.00'

# ==========================================================
# DASHBOARD
# ==========================================================

dash = dashboard

dash.sheet_view.showGridLines = False

# Background
for row in range(1, 60):
    for col in range(1, 11):

        cell = dash.cell(row, col)

        cell.fill = BG_FILL

        dash.column_dimensions[
            get_column_letter(col)
        ].width = 20


# ==========================================================
# TITLE
# ==========================================================

dash.merge_cells("A1:J2")

title = dash["A1"]

title.value = "EQUIPMENT MANAGEMENT DASHBOARD"

title.font = TITLE_FONT

title.alignment = Alignment(
    horizontal="center",
    vertical="center"
)

title.fill = HEADER_FILL


# ==========================================================
# KPI BOXES
# ==========================================================

cards = [

("Total Jobs","B5"),
("Maintenance","E5"),
("Repair","H5"),

("Total Cost","B10"),
("Completed","E10"),
("Pending","H10")

]

for text, pos in cards:

    col = pos[0]

    row = int(pos[1:])

    end_col = get_column_letter(
        ord(col)-64+1
    )

    dash.merge_cells(
        f"{col}{row}:{end_col}{row+2}"
    )

    c = dash[pos]

    c.value = text

    c.font = HEADER_FONT

    c.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    c.fill = HEADER_FILL

    c.border = BORDER


# ==========================================================
# KPI FORMULAS
# ==========================================================

dash["B7"] = f"=COUNTA('{FINANCIAL_SHEET}'!A4:A504)"

dash["E7"] = (
f'=COUNTIF('
f"'{FINANCIAL_SHEET}'!C:C,"
'"Maintenance")'
)

dash["H7"] = (
f'=COUNTIF('
f"'{FINANCIAL_SHEET}'!C:C,"
'"Repair")'
)

dash["B12"] = (
f"=SUM('{FINANCIAL_SHEET}'!L4:L504)"
)

dash["E12"] = (
f'=COUNTIF('
f"'{FINANCIAL_SHEET}'!N:N,"
'"Completed")'
)

dash["H12"] = (
f'=COUNTIF('
f"'{FINANCIAL_SHEET}'!N:N,"
'"Pending")'
)

for ref in [

"B7","E7","H7",

"B12","E12","H12"

]:

    dash[ref].font = BIG_FONT

    dash[ref].alignment = Alignment(
        horizontal="center"
    )

# ==========================================================
# DASHBOARD
# ==========================================================

dash = dashboard

dash.sheet_view.showGridLines = False

# Background
for r in range(1,60):
    for c in range(1,11):
        cell = dash.cell(r,c)
        cell.fill = BG_FILL

for col in range(1,11):
    dash.column_dimensions[get_column_letter(col)].width = 20


# ==========================================================
# TITLE
# ==========================================================

dash.merge_cells("A1:J2")

title = dash["A1"]

title.value = "EQUIPMENT MANAGEMENT DASHBOARD"

title.font = TITLE_FONT

title.fill = HEADER_FILL

title.alignment = Alignment(
    horizontal="center",
    vertical="center"
)


# ==========================================================
# KPI CARDS
# ==========================================================

cards = [

("Maintenance Cost","B5"),

("Repair Cost","E5"),

("Total Cost","H5"),

("Maintenance Jobs","B11"),

("Repair Jobs","E11"),

("Total Jobs","H11")

]

for text,pos in cards:

    col = pos[0]
    row = int(pos[1:])

    dash.merge_cells(
        f"{col}{row}:{get_column_letter(get_column_letter.__globals__['column_index_from_string'](col)+1)}{row+2}"
    )

    c = dash[pos]

    c.value = text

    c.font = HEADER_FONT

    c.fill = HEADER_FILL

    c.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

# ==========================================================
# DASHBOARD
# ==========================================================

dash = wb["Dashboard"]

dash.sheet_view.showGridLines = False

# Column widths
for col in range(1, 11):
    dash.column_dimensions[get_column_letter(col)].width = 18

# Background
bg_fill = PatternFill("solid", fgColor=BACKGROUND)

for r in range(1, 40):
    for c in range(1, 11):
        dash.cell(r, c).fill = bg_fill

# Title
dash.merge_cells("A1:J2")

title = dash["A1"]
title.value = "EQUIPMENT MANAGEMENT DASHBOARD"
title.font = TITLE_FONT
title.fill = HEADER_FILL
title.alignment = Alignment(horizontal="center", vertical="center")

# KPI Boxes
cards = [
    ("Maintenance Jobs", "B5"),
    ("Repair Jobs", "E5"),
    ("Total Cost", "H5"),
]

for text, cell in cards:

    col = cell[0]
    row = int(cell[1:])

    end_col = get_column_letter(get_column_letter(col).index(col)+1 if False else ord(col)-64+1)

    dash.merge_cells(f"{col}{row}:{end_col}{row+2}")

    c = dash[cell]

    c.value = text
    c.font = HEADER_FONT
    c.fill = GROUP_FILL
    c.border = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center")

# KPI formulas

dash["B8"] = '=COUNTIF(\'July Financial Telemetry\'!C:C,"Maintenance")'
dash["E8"] = '=COUNTIF(\'July Financial Telemetry\'!C:C,"Repair")'
dash["H8"] = '=SUM(\'July Financial Telemetry\'!L:L)'

for cell in ("B8","E8","H8"):
    dash[cell].font = BIG_FONT
    dash[cell].alignment = Alignment(horizontal="center")

# Chart

chart = BarChart()

data = Reference(
    financial,
    min_col=12,
    min_row=4,
    max_row=30
)

cats = Reference(
    financial,
    min_col=1,
    min_row=4,
    max_row=30
)

chart.add_data(data)
chart.set_categories(cats)
chart.title = "Monthly Costs"
chart.height = 8
chart.width = 15

dash.add_chart(chart, "B12")

# ==========================================================
# SAVE
# ==========================================================

wb.save(OUTPUT_FILE)

print("="*60)
print("Workbook Created Successfully!")
print(OUTPUT_FILE)
print("="*60)


