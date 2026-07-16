from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
from dateutil.relativedelta import relativedelta

EXCEL_FILE = r"\\apdc-ext01\SGData\SG_Service\Operator\Equipment Status Report.xlsx"

wb = load_workbook(EXCEL_FILE)
ws = wb.active

headers = {
    str(c.value).strip(): c.column
    for c in ws[1]
    if c.value
}

insurance_col = headers["Insurance Expiry"]
lm_col = headers["LM Cert Expiry"]

# Colours
RED = PatternFill(fill_type="solid", start_color="FF0000", end_color="FF0000")
ORANGE = PatternFill(fill_type="solid", start_color="FFC000", end_color="FFC000")
GREEN = PatternFill(fill_type="solid", start_color="00B050", end_color="00B050")

today = datetime.today().date()
warning_date = today + relativedelta(months=2)


def colour_cell(cell):

    if not cell.value:
        return

    try:

        if isinstance(cell.value, datetime):
            expiry = cell.value.date()

        else:
            expiry = datetime.strptime(
                str(cell.value),
                "%d/%m/%Y"
            ).date()

    except Exception:
        print(f"Invalid date: {cell.value}")
        return

    if expiry < today:
        cell.fill = RED

    elif expiry <= warning_date:
        cell.fill = ORANGE

    else:
        cell.fill = GREEN


for row in range(2, ws.max_row + 1):

    insurance_cell = ws.cell(row, insurance_col)
    lm_cell = ws.cell(row, lm_col)

    colour_cell(insurance_cell)
    colour_cell(lm_cell)
    
wb.save(EXCEL_FILE)
wb.close()

print("LM certificate colours updated.")
