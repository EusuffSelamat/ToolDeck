"""
Scan Tanglin Corporation "Cost Sheeting" invoice PDFs (scanned images) via OCR
and append extracted rows to Equipment_Status_Report.xlsx -> sheet "July_Job_Log".

Usage:
    python invoice_to_excel.py <excel_file> <pdf1> [pdf2 ...]

Requires: pytesseract, pdf2image, openpyxl, Pillow + tesseract & poppler installed.
"""
import sys, re, os
import numpy as np
import pdf2image, pytesseract
from pytesseract import Output
from PIL import ImageOps
import openpyxl

# ---------------- Asset type mapping ----------------
SPIDER = {"YC054","YC043","YC052","YC051","YC065","YC083","YC076","YC079","YC077",
          "YC063","YC074","YC106","YC094","YC061"}
BOOM = {"YAB740","YAB743","YAB746","YAB663","YAB733","YAB665",
        "SAB0021","SAB0022","SAB0023","SAB0024"}
SCISSOR = {"YSL662","YSL646","YSL669","YSL671","YSL727","YSL1683","YSL1699","YSL1710",
    "YSL1716","YSL1717","YSL1731","YSL1732","YSL1737","YSL602","YSL603","YSL606",
    "YSL1704","YSL1702","YSL1895","YSL1794","YSL1804","YSL1827","YSL1935","YSL1832",
    "YSL1835","YSL1841","YSL1845","YSL1866","YSL1867","YSL1870","YSL1871","YSL1873",
    "YSL1874","YSL1896","YSL1904","YSL1909","YSL1910","YSL1911","YSL1916","YSL1801",
    "YSL1920","YSL1921","YSL1928","YSL1849","YSL1933","YSL1938","YSL1840","YSL1939",
    "YSL1864","YSL1948","YSL1963","YSL1974","YSL1982","YSL1984","YSL1791","YSL1986",
    "YSL1901","YSL1983","YSL1936","YSL1965","YSL1979","YSL1894","YSL1959","YSL1925",
    "YSL1962","YSL1957"}

def asset_type(plant_no):
    if plant_no in SPIDER: return "Spider Crawler Crane"
    if plant_no in BOOM: return "Boom Lift"
    if plant_no in SCISSOR: return "Scissor Lift"
    if plant_no.startswith("YC"): return "Spider Crawler Crane"
    if plant_no.startswith(("YAB","SAB")): return "Boom Lift"
    if plant_no.startswith("YSL"): return "Scissor Lift"
    return "Unknown"

SERVICE_OPTIONS = [
    ("Monthly Servicing Maintenance with Repair", "Repair"),
    ("Monthly Servicing Maintenance",             "Maintenance"),
    ("Monthly Checking with Top-Up or Repair",    "Maintenance"),
    ("Provide request item(s) / service to Site", "Repair"),
    ("Breakdown Repair @ Site (under Rental)",    "Repair"),
    ("Workshop Repair (Off-Hire)",                "Repair"),
    ("Return to Workshop Repair (Service Exchange)", "Repair"),
]

MONEY = r"([\d,]+\.\d{2})"

def checkbox_fill(img, box):
    region = ImageOps.grayscale(img.crop(box))
    hist = region.histogram()
    total = sum(hist)
    if not total: return 0
    dark = sum(hist[:128])
    return dark / total

def box_interior_fill(img, region):
    crop = ImageOps.grayscale(img.crop(region))
    w, h = crop.size
    px = crop.load()
    EDGE_T = 190
    INK_T = 150
    dark_cols = [sum(1 for y in range(h) if px[x, y] < EDGE_T) for x in range(w)]
    col_thr = max(10, int(h * 0.25))
    cols = [x for x, c in enumerate(dark_cols) if c >= col_thr]
    if not cols:
        return None
    x1, x2 = min(cols), max(cols)
    bw = x2 - x1
    if not (12 <= bw <= 60):
        return None
    ys = [y for y in range(h) if px[x1, y] < EDGE_T or px[x2, y] < EDGE_T]
    if not ys:
        return None
    y1, y2 = min(ys), max(ys)
    bh = y2 - y1
    if not (12 <= bh <= 60):
        return None
    ix1, ix2 = x1 + int(bw * 0.25), x2 - int(bw * 0.25)
    iy1, iy2 = y1 + int(bh * 0.25), y2 - int(bh * 0.25)
    if ix2 <= ix1 or iy2 <= iy1:
        return None
    total = (ix2 - ix1) * (iy2 - iy1)
    dark = sum(1 for x in range(ix1, ix2) for y in range(iy1, iy2) if px[x, y] < INK_T)
    return dark / total

def box_tick_density(img, x0, x1, yc):
    import numpy as np
    from scipy import ndimage
    y0, y1 = yc - 38, yc + 44
    g = np.asarray(ImageOps.grayscale(img.crop((int(x0), int(y0), int(x1), int(y1)))))
    b = g < 160
    lab, _n = ndimage.label(b, structure=np.ones((3, 3)))
    best = None
    for sl in ndimage.find_objects(lab):
        h = sl[0].stop - sl[0].start
        w = sl[1].stop - sl[1].start
        if 18 <= h <= 48 and 18 <= w <= 48 and 0.6 <= w / h <= 1.7:
            comp = b[sl]
            if comp[0, :].mean() > 0.5 or comp[-1, :].mean() > 0.5:
                if best is None or h * w > best[0]:
                    best = (h * w, sl)
    if best is None:
        return None
    sl = best[1]
    inner = b[sl[0].start + 5:sl[0].stop - 5, sl[1].start + 5:sl[1].stop - 5]
    return float(inner.mean()) if inner.size else None

def detect_work_type(img):
    data = pytesseract.image_to_data(img, output_type=Output.DICT)
    words = [(data['text'][i].strip(), data['left'][i], data['top'][i],
              data['width'][i], data['height'][i])
             for i in range(len(data['text'])) if data['text'][i].strip()]

    def same_line(w, token):
        return any(token.lower() in o[0].lower() and abs(o[2] - w[2]) < 14
                   and o[1] > w[1] for o in words)

    label_defs = [
        ("Monthly",   ("Servicing", "with"), None,    "Repair"),
        ("Monthly",   ("Servicing",),        "with",  "Maintenance"),
        ("Monthly",   ("Checking",),         None,    "Maintenance"),
        ("Provide",   ("request",),          None,    "Repair"),
        ("Breakdown", (),                    None,    "Repair"),
        ("Workshop",  ("(Off",),             None,    "Repair"),
        ("Return",    ("Exchange",),         None,    "Repair"),
    ]
    results = []
    used = []
    for anchor, req, excl, wtype in label_defs:
        for w in words:
            if anchor.lower() not in w[0].lower():
                continue
            if req and not all(same_line(w, r) for r in req):
                continue
            if excl and same_line(w, excl):
                continue
            if any(abs(w[2] - t) < 12 and abs(w[1] - x) < 120 for t, x in used):
                continue
            used.append((w[2], w[1]))
            t, x, y, wd, h = w
            merged = not t[0].isalnum()
            if merged:
                win = (x - 6, x + 60)
            else:
                win = (max(0, x - 90), x - 4)
            s = box_tick_density(img, win[0], win[1], y + h // 2)
            if s is not None:
                results.append((s, wtype))
            break
    results.sort(reverse=True)
    if results and results[0][0] >= 0.02:
        return results[0][1]
    return None

def ocr_lines(img):
    d = pytesseract.image_to_data(img, output_type=Output.DICT)
    words = []
    for i in range(len(d['text'])):
        t = d['text'][i].strip()
        if t:
            words.append((d['top'][i], d['left'][i], t))
    words.sort()
    grouped = []
    for top, left, t in words:
        if grouped and abs(grouped[-1][0] - top) < 18:
            grouped[-1][1].append((left, t))
        else:
            grouped.append((top, [(left, t)]))
    return [' '.join(t for l, t in sorted(ws)) for top, ws in grouped]

def extract_section_totals(lines):
    def section_bounds(start_kws, end_kws):
        if isinstance(start_kws, str):
            start_kws = [start_kws]
        for kw in start_kws:
            s = e = None
            for i, l in enumerate(lines):
                if s is None and kw.lower() in l.lower():
                    s = i
                elif s is not None and any(k.lower() in l.lower() for k in end_kws):
                    e = i; break
            if s is not None:
                return (s, e if e is not None else len(lines))
        return (None, len(lines))

    out = {}
    s, e = section_bounds(["Spare Parts", "External Services",
                           "Unit Rate", "Category"], ["Transmission Oil"])
    items = []
    if s is not None:
        for l in lines[s+1:e]:
            m = re.search(r"(\d+)\s+(p[ce]s?)\b", l, re.I)
            moneys = re.findall(MONEY, l)
            if m and moneys:
                qty = int(m.group(1))
                if qty > 1 and not m.group(2).lower().endswith("s"):
                    print(f"    note: qty {qty} next to singular 'pc' - correcting to 1 ({l.strip()[:60]})")
                    qty = 1
                desc = l[:m.start()]
                desc = re.sub(r"^\W*\d{1,2}\b\W*", "", desc)
                desc = re.sub(r"^(?:[^\s]{1,2}\s+){0,2}(?=[A-Z])", "", desc.strip())
                desc = re.sub(r"\(\s+", "(", desc)
                desc = re.sub(r"\s+\)", ")", desc)
                desc = re.sub(r"\s+", " ", desc).strip(" .:-|_")
                if not desc:
                    desc = "(unreadable)"
                items.append((desc, qty, float(moneys[-1].replace(",",""))))
    out["Spare Parts & External Services"] = items

    MATERIAL_NAMES = ["Transmission Oil", "Engine Oil", "Hydraulic Oil", "Brake Oil",
                      "Compressor Oil", "Washing & Cement Hacking", "Washing & Painting",
                      "Washing & Welding", "Welding", "Diesel Tank Cleaning", "Battery Water"]
    s, e = section_bounds("Transmission Oil", ["Hrs", "Labours"])
    items = []
    if s is not None:
        for l in lines[s:e]:
            m = re.search(r"(\d+(?:\.\d+)?)\s+(?:ltrs?|itrs?|1trs?|job)\b", l, re.I)
            moneys = re.findall(MONEY, l)
            if m and len(moneys) >= 2:
                low = l.lower()
                desc = next((n for n in MATERIAL_NAMES
                             if all(w in low for w in n.lower().split() if w != "&")), None)
                if desc is None:
                    desc = re.sub(r"^\s*\d+\s*\W*", "", l[:m.start(1)]).strip(" .:-")
                items.append((desc, float(m.group(1)), float(moneys[-1].replace(",",""))))
    out["Materials"] = items

    s, e = section_bounds("Hrs", ["Transportation", "Total"])
    qty = 0; cost = 0.0
    if s is not None:
        for l in lines[s:e]:
            if re.search(r"\btrip\b", l, re.I):
                continue
            l2 = l.replace("S$", "$").replace("$$", "$")
            m = re.search(r"(?<![\d.,])(\d+)(?:\s+(\d+))?\s+\$?\s*[\d,]+\.\d{2}", l2)
            moneys = re.findall(MONEY, l2)
            if m and len(moneys) >= 2:
                hrs = int(m.group(1)) + (int(m.group(2))/60 if m.group(2) else 0)
                qty += hrs; cost += float(moneys[-1].replace(",",""))
    out["Labours"] = (qty, cost)

    qty = 0; cost = 0.0
    if True:
        for l in lines:
            m = re.search(r"(\d+)\s+trip\b", l, re.I)
            moneys = re.findall(MONEY, l)
            if m and len(moneys) >= 2:
                qty += int(m.group(1)); cost += float(moneys[-1].replace(",",""))
    out["Transportation"] = (qty, cost)
    return out

def parse_page(img, filename):
    lines = ocr_lines(img)
    text = "\n".join(lines)
    plant = date = None
    m = re.search(r"Plant Number:\s*([A-Z]{2,3}\d+)", text)
    if m: plant = m.group(1)
    m = re.search(r"Date\s*:?\s*(\d{1,2}/\d{1,2}/\d{4})", text)
    if m: date = m.group(1)

    work = detect_work_type(img)
    if work is None:
        low = text.lower()
        repair_score = sum(k in low for k in
            ["replace", "spoiled", "leak", "not functioning", "breakdown",
             "complaint", "repair it", "order a new", "change it"])
        maint_score = sum(k in low for k in
            ["monthly checking", "perform monthly", "function test ok",
             "performance test ok", "good condition"])
        work = "Repair" if repair_score >= maint_score else "Maintenance"

    if date:
        d, mo, y = date.split("/")
        date = f"{int(d):02d}/{int(mo):02d}/{y}"

    sections = extract_section_totals(lines)
    sp_cost = sum(c for _, _, c in sections["Spare Parts & External Services"])
    mat_cost = sum(c for _, _, c in sections["Materials"])
    lab = sections["Labours"]
    tr = sections["Transportation"]
    total = sp_cost + mat_cost + lab[1] + tr[1]

    printed = None
    m = re.search(r"Total\s*:?.{0,20}?([\d,]+\.\d{2})", text)
    if m:
        printed = float(m.group(1).replace(",", ""))

    return {"date": date, "plant": plant, "asset": asset_type(plant or ""),
            "work": work, "sections": sections, "total": total, "file": filename,
            "printed_total": printed, "ocr_lines": lines}

from openpyxl.utils import get_column_letter
from copy import copy as _copy
from datetime import datetime

LAST_TEMPLATE_ROW = 500


def _fix_total_formulas(ws, header_row, start_row, end_row):
    """Fix Total column formulas to use correct column references after Type insertion.
    
    After Type columns are inserted, the column layout is:
    - H (8): Spare Parts Cost (total cost for line item)
    - K (11): Materials Cost (total cost for line item)
    - M (13): Labour Cost (total cost)
    - O (15): Transport Cost (total cost)
    """
    total_col = None
    for c in range(1, 25):
        val = ws.cell(header_row, c).value
        if val and "Total" in str(val).strip():
            total_col = c
            break
    
    if total_col is None:
        print("    Warning: Total column header not found, cannot fix formulas")
        return False
    
    for row in range(start_row, end_row + 1):
        formula = (
            f'=IFERROR(H{row},0)'
            f'+IFERROR(K{row},0)'
            f'+IFERROR(M{row},0)'
            f'+IFERROR(O{row},0)'
        )
        ws.cell(row, total_col).value = formula
    
    return True


def _find_last_data_row(ws, header_row):
    """Find the last row that contains data (checks S/N and Date columns)."""
    last = header_row
    r = header_row + 1
    while r < LAST_TEMPLATE_ROW and (ws.cell(r, 1).value not in (None, "") or 
                                      ws.cell(r, 2).value not in (None, "")):
        last = r
        r += 1
    return last


def ensure_type_columns(ws, header_row):
    """Insert 'Type' columns before the Qty of Spare Parts (F) and Materials,
    copying the template's theme. Safe to call repeatedly - also repairs the
    theme of already-inserted columns AND fixes Total column formulas."""
    inserted = False
    if ws.cell(header_row, 6).value != "Type":
        ws.insert_cols(8)
        ws.insert_cols(6)
        inserted = True

    last = _find_last_data_row(ws, header_row)
    br = header_row - 1

    for new_col, src_col in ((6, 7), (9, 10)):
        for row in range(1, last + 1):
            if row == br:
                continue
            ws.cell(row, new_col)._style = _copy(ws.cell(row, src_col)._style)
        ws.column_dimensions[get_column_letter(new_col)].width = 20
        ws.cell(header_row, new_col, "Type")

    if inserted:
        for m in [m for m in list(ws.merged_cells.ranges) if m.min_row == br]:
            ws.unmerge_cells(str(m))
        for c in range(6, 16):
            if ws.cell(br, c).value not in (None, "") and c not in (6, 9, 12, 14):
                ws.cell(br, c).value = None
    banner_style = ws.cell(br, 12)._style
    banners = [(6, 8, "Spare Parts & External Services"), (9, 11, "Materials"),
               (12, 13, "Labours"), (14, 15, "Transportation")]
    existing = {(m.min_col, m.max_col) for m in ws.merged_cells.ranges if m.min_row == br}
    for c1, c2, label in banners:
        for c in range(c1, c2 + 1):
            ws.cell(br, c)._style = _copy(banner_style)
        ws.cell(br, c1, label)
        if (c1, c2) not in existing:
            for m in [m for m in list(ws.merged_cells.ranges)
                      if m.min_row == br and not (m.max_col < c1 or m.min_col > c2)]:
                ws.unmerge_cells(str(m))
            ws.merge_cells(start_row=br, start_column=c1, end_row=br, end_column=c2)
    
    if last > header_row:
        _fix_total_formulas(ws, header_row, header_row + 1, last)


def sort_by_date(ws, header_row):
    """Sort all data rows by Date column (oldest to newest) and re-number S/N.
    
    Handles dates in DD/MM/YYYY string format and datetime objects.
    Rows with no date are placed at the end.
    Preserves hyperlinks in the File column.
    """
    from openpyxl.styles import Font
    link_font = Font(color="0563C1", underline="single")
    
    last_row = _find_last_data_row(ws, header_row)
    
    if last_row <= header_row:
        print("    No data to sort")
        return 0
    
    num_rows = last_row - header_row
    print(f"    Sorting {num_rows} rows by date (oldest to newest)...")
    
    # Collect all row data
    rows_data = []
    for r in range(header_row + 1, last_row + 1):
        date_val = ws.cell(r, 2).value
        parsed_date = None
        
        if date_val:
            if isinstance(date_val, datetime):
                parsed_date = date_val
            elif isinstance(date_val, str):
                # Try DD/MM/YYYY format first
                for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y"):
                    try:
                        parsed_date = datetime.strptime(date_val, fmt)
                        break
                    except ValueError:
                        continue
        
        # Collect all cell values and hyperlinks
        row_vals = {}
        hyperlink = None
        max_col = 20  # Cover all used columns
        for c in range(1, max_col + 1):
            row_vals[c] = ws.cell(r, c).value
            if c == 17:  # File column
                hyperlink = ws.cell(r, c).hyperlink
        
        rows_data.append({
            'parsed_date': parsed_date,
            'vals': row_vals,
            'hyperlink': hyperlink,
            'original_row': r,
        })
    
    # Count dates that were successfully parsed
    parsed_count = sum(1 for r in rows_data if r['parsed_date'] is not None)
    unparsed_count = num_rows - parsed_count
    
    # Sort by parsed date (None dates go to end)
    rows_data.sort(key=lambda x: (x['parsed_date'] is None, x['parsed_date'] or datetime.max))
    
    # Clear the data area (values and hyperlinks only, keep styles)
    for r in range(header_row + 1, last_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.value = None
            if c == 17:
                cell.hyperlink = None
    
    # Write back sorted data
    for i, row_data in enumerate(rows_data):
        r = header_row + 1 + i
        for c, v in row_data['vals'].items():
            if c == 1:
                ws.cell(r, c, i + 1)  # Re-number S/N starting from 1
            else:
                ws.cell(r, c, v)
        
        # Restore hyperlink
        if row_data['hyperlink']:
            fcell = ws.cell(r, 17)
            fcell.hyperlink = row_data['hyperlink']
            fcell.font = link_font
    
    # Fix Total formulas for all rows after sorting
    _fix_total_formulas(ws, header_row, header_row + 1, header_row + num_rows)
    
    print(f"    Sorted: {parsed_count} rows with valid dates" + 
          (f", {unparsed_count} rows without dates placed at end" if unparsed_count > 0 else ""))
    
    return num_rows


def append_to_excel(excel_path, rows, sheet="July_Job_Log"):
    from openpyxl.styles import Font
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet]
    header_row = None
    for r in range(1, 30):
        if ws.cell(r, 1).value == "S/N":
            header_row = r; break
    if header_row is None:
        raise RuntimeError("Header row with 'S/N' not found")

    ensure_type_columns(ws, header_row)

    # Find Total column for writing formulas to new rows
    total_col = None
    for c in range(1, 25):
        val = ws.cell(header_row, c).value
        if val and "Total" in str(val).strip():
            total_col = c
            break

    # first empty row (Date column)
    r = header_row + 1
    while ws.cell(r, 2).value not in (None, ""):
        r += 1

    link_font = Font(color="0563C1", underline="single")
    first_new_row = r

    for rec in rows:
        sp_items = rec["sections"]["Spare Parts & External Services"]
        mat_items = rec["sections"]["Materials"]
        lab = rec["sections"]["Labours"]
        tr = rec["sections"]["Transportation"]
        n_lines = max(len(sp_items), len(mat_items), 1)

        for i in range(n_lines):
            sp = sp_items[i] if i < len(sp_items) else None
            mat = mat_items[i] if i < len(mat_items) else None
            first = (i == 0)
            prev_sn = ws.cell(r - 1, 1).value
            sn = prev_sn + 1 if isinstance(prev_sn, (int, float)) else 1
            vals = {
                1: sn, 2: rec["date"], 3: rec["plant"], 4: rec["asset"], 5: rec["work"],
                6: sp[0] if sp else None, 7: sp[1] if sp else None, 8: sp[2] if sp else None,
                9: mat[0] if mat else None, 10: mat[1] if mat else None, 11: mat[2] if mat else None,
                12: (lab[0] or None) if first else None, 13: (lab[1] or None) if first else None,
                14: (tr[0] or None) if first else None, 15: (tr[1] or None) if first else None,
            }
            for c, v in vals.items():
                cur = ws.cell(r, c).value
                if c == 1 and cur not in (None, ""):
                    continue
                if v is not None:
                    ws.cell(r, c, v)
            fcell = ws.cell(r, 17)
            fcell.value = rec["date"] or rec["file"]
            if rec.get("path"):
                fcell.hyperlink = "file:///" + os.path.abspath(rec["path"]).replace("\\", "/")
                fcell.font = link_font
            r += 1
    
    # Write Total formulas for newly added rows
    if total_col is not None and r > first_new_row:
        _fix_total_formulas(ws, header_row, first_new_row, r - 1)
    
    # --- SORT ALL DATA BY DATE (OLDEST TO NEWEST) ---
    print("\n  Sorting data by date...")
    sort_by_date(ws, header_row)
    
    wb.save(excel_path)
    print(f"  Data sorted and saved.\n")

TYPE_FOLDERS = {"Spider Crane": "Spider Crawler Crane",
                "Boom Lift": "Boom Lift",
                "Scissor Lift": "Scissor Lift"}

def find_tesseract():
    import shutil
    if shutil.which("tesseract"):
        return
    for cand in [r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                 r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
                 os.path.expanduser(r"~\AppData\Local\Programs\Tesseract-OCR\tesseract.exe"),
                 os.path.expanduser(r"~\AppData\Local\Tesseract-OCR\tesseract.exe")]:
        if os.path.exists(cand):
            pytesseract.pytesseract.tesseract_cmd = cand
            return
    import tkinter as tk
    from tkinter import filedialog, messagebox
    r = tk.Tk(); r.withdraw()
    messagebox.showinfo("Tesseract needed",
        "Tesseract OCR was not found on this PC.\n\n"
        "Please select tesseract.exe.\n\n"
        "If you haven't installed it yet, download it from:\n"
        "github.com/UB-Mannheim/tesseract/wiki")
    exe = filedialog.askopenfilename(title="Select tesseract.exe",
                                     filetypes=[("tesseract.exe", "tesseract.exe")])
    r.destroy()
    if exe:
        pytesseract.pytesseract.tesseract_cmd = exe
    else:
        raise SystemExit("Tesseract not found - cannot OCR PDFs. Install Tesseract and try again.")

def find_poppler():
    import shutil, glob
    if shutil.which("pdfinfo"):
        return None
    candidates = []
    for base in [r"C:\poppler", r"C:\Program Files\poppler", r"C:\Program Files (x86)\poppler",
                 os.path.expanduser(r"~\poppler"),
                 os.path.expanduser(r"~\Downloads\poppler"),
                 os.path.expanduser(r"~\Desktop\poppler")]:
        candidates += glob.glob(base + "*")
    for cand in candidates:
        for root_dir, dirs, files in os.walk(cand):
            if "pdfinfo.exe" in files:
                return root_dir
    import tkinter as tk
    from tkinter import filedialog, messagebox
    r = tk.Tk(); r.withdraw()
    messagebox.showinfo("Poppler needed",
        "Poppler was not found on this PC.\n\n"
        "Please select the Poppler 'bin' folder (the one containing pdfinfo.exe).\n\n"
        "If you haven't installed it yet, download it from:\n"
        "github.com/oschwartz10612/poppler-windows/releases\n"
        "and extract the zip anywhere (e.g. C:\\poppler).")
    folder = filedialog.askdirectory(title="Select Poppler 'bin' folder (contains pdfinfo.exe)")
    r.destroy()
    if folder and os.path.exists(os.path.join(folder, "pdfinfo.exe")):
        return folder
    raise SystemExit("Poppler not found - cannot convert PDFs. Install Poppler and try again.")

def scan_invoice_folder(invoice_folder):
    found = []
    for type_name in sorted(os.listdir(invoice_folder)):
        type_path = os.path.join(invoice_folder, type_name)
        if not os.path.isdir(type_path):
            continue
        asset = TYPE_FOLDERS.get(type_name.strip(), None)
        for plant_name in sorted(os.listdir(type_path)):
            plant_path = os.path.join(type_path, plant_name)
            if not os.path.isdir(plant_path):
                continue
            plant_no = plant_name.strip().upper()
            for root_dir, dirs, files in os.walk(plant_path):
                for f in sorted(files):
                    if f.lower().endswith(".pdf"):
                        found.append((os.path.join(root_dir, f), plant_no, asset))
    return found

def main():
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk()
    root.withdraw()

    print("Select the Excel file (Equipment_Status_Report.xlsx)...")
    excel = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
    )
    if not excel:
        print("No Excel file selected. Exiting.")
        return

    print("Where is the Invoice folder?")
    invoice_folder = filedialog.askdirectory(title="Where is the Invoice folder?")
    root.destroy()
    if not invoice_folder:
        print("No folder selected. Exiting.")
        return

    poppler_path = find_poppler()
    find_tesseract()

    pdf_list = scan_invoice_folder(invoice_folder)
    if not pdf_list:
        print(f"No PDFs found under {invoice_folder}\\<Type>\\<Plant No.>\\")
        return

    print(f"\nExcel:  {excel}")
    print(f"Folder: {invoice_folder}")
    print(f"Found:  {len(pdf_list)} PDF(s)\n")

    rows = []
    errors = []
    for pdf, folder_plant, folder_asset in pdf_list:
        try:
            pages = pdf2image.convert_from_path(pdf, dpi=300, poppler_path=poppler_path)
            for i, img in enumerate(pages):
                fname = os.path.basename(pdf) + (f" p{i+1}" if len(pages) > 1 else "")
                rec = parse_page(img, fname)
                rec["path"] = os.path.abspath(pdf)
                if rec["plant"] and rec["plant"] != folder_plant:
                    print(f"  WARNING: {fname} OCR read plant '{rec['plant']}' "
                          f"but file is in folder '{folder_plant}' - using folder value")
                rec["plant"] = folder_plant
                rec["asset"] = folder_asset or asset_type(folder_plant)
                print(f"{rec['file']}: {rec['date']} {rec['plant']} {rec['asset']} {rec['work']} "
                      f"total=${rec['total']:.2f} sections={rec['sections']}")
                pt = rec.get("printed_total")
                if pt is not None and abs(pt - rec["total"]) > 0.01:
                    dbg = os.path.join(os.path.dirname(os.path.abspath(excel)),
                                       "_ocr_debug_" + os.path.splitext(os.path.basename(pdf))[0] + ".txt")
                    with open(dbg, "w", encoding="utf-8") as fh:
                        fh.write(f"File: {pdf}\nExtracted total: {rec['total']}\n"
                                 f"Printed total on invoice: {pt}\n"
                                 f"Sections: {rec['sections']}\n\n--- OCR lines ---\n")
                        fh.write("\n".join(rec["ocr_lines"]))
                    print(f"  WARNING: extracted ${rec['total']:.2f} but invoice says ${pt:.2f}"
                          f" - OCR dump saved to {dbg}")
                rows.append(rec)
        except Exception as e:
            errors.append((pdf, str(e)))
            print(f"  ERROR reading {pdf}: {e}")

    if rows:
        append_to_excel(excel, rows)
        print(f"Appended {len(rows)} rows to {excel} -> July_Job_Log")
        print("(All data sorted by date: oldest first at S/N 1)")
    if errors:
        print(f"\n{len(errors)} file(s) failed:")
        for p, e in errors:
            print(f"  {p}: {e}")

if __name__ == "__main__":
    main()
