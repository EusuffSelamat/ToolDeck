import os
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

# ==========================================================
# FILE LOCATIONS
# ==========================================================

EXCEL_FILE = r"C:\Users\eusuff.binselamat\Downloads\Equipment Status Report.xlsx"

SERVICE_REPORT_ROOT = r"\\apdc-ext01\SGData\SG_Service\Operator\Invoice"

# ==========================================================
# OPEN WORKBOOK
# ==========================================================

wb = load_workbook(EXCEL_FILE)
ws = wb.active

# ==========================================================
# FIND COLUMN HEADERS
# ==========================================================

headers = {}

for cell in ws[1]:
    if cell.value:
        headers[str(cell.value).strip()] = cell.column

plant_col = headers["Plant No."]
invoice_col = headers["Invoice Receipts"]

# ==========================================================
# DETERMINE WHICH MAIN FOLDER TO USE
# ==========================================================

def get_category(plant):

    if plant.startswith(("YAB", "SAB")):
        return "Boom Lift"

    elif plant.startswith("YSL"):
        return "Scissor Lift"

    else:
        return "Spider Crane"


# ==========================================================
# FIND THE ACTUAL FOLDER
# (allows folders like "YC054 - Service Reports")
# ==========================================================

from datetime import datetime
import os
import re

def get_latest_report(folder):
    """
    Returns:
        display_name (without .pdf)
        full_path
    """

    latest_date = None
    latest_file = None

    for file in os.listdir(folder):

        if not file.lower().endswith(".pdf"):
            continue

        # Find dates like 13.05.2027 or 13-05-2027
        m = re.search(r"(\d{2})[.-](\d{2})[.-](\d{4})", file)

        if not m:
            continue

        try:
            file_date = datetime.strptime(
                m.group(0).replace("-", "."),
                "%d.%m.%Y"
            )

            if latest_date is None or file_date > latest_date:
                latest_date = file_date
                latest_file = file

        except ValueError:
            pass

    if latest_file:
        display_name = latest_date.strftime("%d/%m/%Y")

        return (
            display_name,
            os.path.join(folder, latest_file)
            )

    return ("N/A", folder)

def find_folder(plant):

    category = get_category(plant)

    base = os.path.join(
        SERVICE_REPORT_ROOT,
        category
    )

    if not os.path.isdir(base):
        return None

    # Exact match first
    exact = os.path.join(base, plant)

    if os.path.isdir(exact):
        return exact

    # Otherwise look for folders beginning with Plant No.
    for folder in os.listdir(base):

        full = os.path.join(base, folder)

        if os.path.isdir(full):

            if folder.upper().startswith(plant.upper()):
                return full

    return None


# ==========================================================
# MAIN LOOP
# ==========================================================

for row in range(2, ws.max_row + 1):

    plant = str(ws.cell(row=row, column=plant_col).value or "").strip()

    plant = plant.split("(")[0].strip()

    if plant == "":
        continue

    invoice_cell = ws.cell(row=row, column=invoice_col)

    # Skip rows already processed
    if invoice_cell.value not in (None, ""):
        print(f"Skipping {plant}")
        continue

    print("=" * 60)
    print(f"Processing {plant}")

    folder = find_folder(plant)

    if folder is None:

        print("Folder not found.")
        continue

    print(folder)

    display_name, _ = get_latest_report(folder)

    invoice_cell.value = display_name
    invoice_cell.hyperlink = folder

    invoice_cell.font = Font(
    color="0563C1",
    underline="single"
    )

    print("✓ Saved")

# ==========================================================
# SAVE
# ==========================================================

wb.save(EXCEL_FILE)
wb.close()

print()
print("=" * 60)
print("Finished!")
print("=" * 60)
