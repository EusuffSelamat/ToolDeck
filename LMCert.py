import os
import re
import fitz                    # PyMuPDF
import pytesseract

from pdf2image import convert_from_path

from openpyxl import load_workbook
from openpyxl.styles import Font

EXCEL_FILE = r"C:\Users\eusuff.binselamat\Downloads\Equipment Status Report.xlsx"
LM_ROOT = r"\\apdc-ext01\SGData\SG_Service\Operator\LM Cert"

# Path to your Tesseract installation
pytesseract.pytesseract.tesseract_cmd = (
    r"C:\Users\eusuff.binselamat\AppData\Local\Programs\Tesseract-OCR\tesseract.exe"
)

wb=load_workbook(EXCEL_FILE)
ws=wb.active
headers={str(c.value).strip():c.column for c in ws[1] if c.value}

plant_col=headers["Plant No."]
reg_col=headers["LE Registration No."]
expiry_col=headers["LM Cert Expiry"]
yom_col=headers["YOM"]

def category(p):
    if p.startswith(("YAB","SAB")): return "Boom Lift"
    if p.startswith("YSL"): return "Scissor Lift"
    return "Spider Crane"

def folder_for(plant):
    base=os.path.join(LM_ROOT,category(plant),plant)
    return base if os.path.isdir(base) else None

def find_pdf(folder):
    for root,_,files in os.walk(folder):
        for f in files:
            if f.lower().endswith(".pdf"):
                return os.path.join(root,f)
    return None

def clean(t):
    t=t.replace("\r","\n")
    t=re.sub(r"[ \t]+"," ",t)
    t=re.sub(r"\n+","\n",t)
    return t

def after(text, labels):

    for lab in labels:

        # Value on the same line
        m = re.search(
            rf"{lab}\s*:?\s*([^\n]+)",
            text,
            re.I
        )

        if m:
            return m.group(1).strip()

        # Value on the next line
        m = re.search(
            rf"{lab}\s*:?\s*\n\s*([^\n]+)",
            text,
            re.I
        )

        if m:
            return m.group(1).strip()

    return ""

def parse(pdf):

    txt = ""

    # -----------------------------
    # First attempt: PyMuPDF
    # -----------------------------
    doc = fitz.open(pdf)

    for page in doc:
        txt += page.get_text("text") + "\n"

    doc.close()

    txt = clean(txt)

    # -----------------------------
    # OCR fallback
    # -----------------------------
    if not txt.strip():

        print("OCR:", os.path.basename(pdf))

        pages = convert_from_path(
            pdf,
            dpi=300,
            poppler_path=r"C:\poppler\Library\bin"
        )
        for img in pages:

            txt += pytesseract.image_to_string(
                img,
                lang="eng"
            )

            txt += "\n"

        txt = clean(txt)

    # -----------------------------
    # Registration Number
    # -----------------------------
    m = re.search(
        r"\b(?:LP|LM)\d{5,}[A-Z]?\b",
        txt,
        re.I
    )

    reg = m.group(0).upper() if m else ""

    # -----------------------------
    # Expiry Date
    # -----------------------------
    exp = after(
        txt,
        [
            r"Certificate\s*Expiry\s*Date",
            r"Expiry\s*Date",
            r"Certificate\s*Expiry"
        ]
    )

    if not exp:

        m = re.search(
            r"\b\d{2}/\d{2}/\d{4}\b",
            txt
        )

        if m:
            exp = m.group(0)

    # -----------------------------
    # Year of Manufacture
    # -----------------------------
    yom = after(
        txt,
        [
            r"YEAR\s*OF\s*MFG",
            r"YEAR\s*OF\s*MANUFACTURE",
            r"Manufacture\s*Year"
        ]
    )

    m = re.search(
        r"\b(19|20)\d{2}\b",
        yom
    )

    if not m:

        m = re.search(
            r"\b(19|20)\d{2}\b",
            txt
        )

    yom = m.group(0) if m else ""

    return reg, exp, yom

for r in range(2,ws.max_row+1):
    plant=str(ws.cell(r,plant_col).value or "").split("(")[0].strip()
    if not plant: continue
    reg_cell = ws.cell(r, reg_col)

    if (
        reg_cell.value not in (None, "")
        and reg_cell.hyperlink is not None
        and ws.cell(r, expiry_col).value not in (None, "")
        and ws.cell(r, yom_col).value not in (None, "")
    ):
        print("Skipping", plant)
        continue
    fld=folder_for(plant)
    if not fld:
        print("Missing folder",plant); continue
    pdf=find_pdf(fld)
    if not pdf:
        print("Missing PDF",plant); continue
    try:
        reg, exp, yom = parse(pdf)

        # ----------------------------
        # LE Registration No. (Clickable)
        # ----------------------------
        reg_cell = ws.cell(r, reg_col)

        # Show the registration number
        reg_cell.value = reg

        # Create a relative hyperlink to the PDF
        # Link directly to the network PDF
        reg_cell.hyperlink = pdf

        # Make it look like a hyperlink
        reg_cell.font = Font(
            color="0563C1",
            underline="single"
        )

        # ----------------------------
        # Certificate Expiry
        # ----------------------------
        ws.cell(r, expiry_col).value = exp

        # ----------------------------
        # Year of Manufacture
        # ----------------------------
        ws.cell(r, yom_col).value = yom
        print(f"{plant}: {reg} | {exp} | {yom}")
    except Exception as e:
        print("Error",plant,e)

wb.save(EXCEL_FILE)
wb.close()
print("Done.")
