import os
import re
import fitz  # PyMuPDF
import pytesseract
from pdf2image import convert_from_path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from pytesseract import Output

# =====================================================
# CONFIGURATION
# =====================================================

EXCEL_FILE = r"C:\Users\eusuff.binselamat\Downloads\Equipment Status Report.xlsx"
INVOICE_ROOT = r"C:\Users\eusuff.binselamat\Downloads\Invoice"
POPPLER_PATH = r"C:\poppler\Library\bin"

pytesseract.pytesseract.tesseract_cmd = (
    r"C:\Users\eusuff.binselamat\AppData\Local\Programs\Tesseract-OCR\tesseract.exe"
)

# =====================================================
# ASSET TYPES
# =====================================================

SPIDER = {
    "YC054","YC043","YC052","YC051","YC065","YC083",
    "YC076","YCO79","YC077","YC063","YC074","YC106",
    "YC094","YC061"
}

BOOM = {
    "YAB740","YAB743","YAB746","YAB663",
    "YAB733","YAB665",
    "SAB0021","SAB0022","SAB0023","SAB0024"
}

SCISSOR = {
    "YSL662","YSL646","YSL669","YSL671","YSL727",
    "YSL1683","YSL1699","YSL1710","YSL1716","YSL1717",
    "YSL1731","YSL1732","YSL1737","YSL602","YSL603",
    "YSL606","YSL1704","YSL1702","YSL1895","YSL1794",
    "YSL1804","YSL1827","YSL1935","YSL1832","YSL1835",
    "YSL1841","YSL1845","YSL1866","YSL1867","YSL1870",
    "YSL1871","YSL1873","YSL1874","YSL1896","YSL1904",
    "YSL1909","YSL1910","YSL1911","YSL1916","YSL1801",
    "YSL1920","YSL1921","YSL1928","YSL1849","YSL1933",
    "YSL1938","YSL1840","YSL1939","YSL1864","YSL1948",
    "YSL1963","YSL1974","YSL1982","YSL1984","YSL1791",
    "YSL1986","YSL1901","YSL1983","YSL1936","YSL1965",
    "YSL1979","YSL1894","YSL1959","YSL1925","YSL1962",
    "YSL1957"
}

def asset_type(plant):
    if plant in SPIDER:
        return "Spider Crawler Crane"
    if plant in BOOM:
        return "Boom Lift"
    if plant in SCISSOR:
        return "Scissor Lift"
    return ""

# =====================================================
# TEXT & MATH HELPERS
# =====================================================

def clean(text):
    text = text.replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n+", "\n", text)
    return text.strip()

def is_number(txt):
    cleaned = txt.replace(",", "").replace("$", "").replace("S$", "").strip()
    return re.fullmatch(r"\d+(?:\.\d+)?", cleaned) is not None

def money(value):
    if value is None:
        return 0.0
    value = str(value).replace(",", "").replace("$", "").replace("S$", "").strip()
    try:
        return float(value)
    except ValueError:
        return 0.0

# =====================================================
# HYBRID WORD & LINE EXTRACTOR
# =====================================================

def group_words_into_lines(words, tolerance):
    if not words:
        return []
    
    words.sort(key=lambda w: (w["page"], w["y"], w["x"]))
    lines = []
    current_line = []
    current_y = None
    current_page = None
    
    for w in words:
        if current_page is None or w["page"] != current_page:
            if current_line:
                current_line.sort(key=lambda x: x["x"])
                lines.append(current_line)
            current_line = [w]
            current_y = w["y"]
            current_page = w["page"]
            continue
            
        if current_y is None:
            current_y = w["y"]
            current_line.append(w)
        elif abs(w["y"] - current_y) <= tolerance:
            current_line.append(w)
        else:
            current_line.sort(key=lambda x: x["x"])
            lines.append(current_line)
            current_line = [w]
            current_y = w["y"]
            
    if current_line:
        current_line.sort(key=lambda x: x["x"])
        lines.append(current_line)
        
    return lines

def get_pdf_words_and_lines(pdf_path):
    """Extracts words digitally first, falls back to Tesseract OCR if empty."""
    doc = fitz.open(pdf_path)
    is_scanned = True
    
    for page in doc:
        if page.get_text().strip():
            is_scanned = False
            break
            
    lines = []
    
    if not is_scanned:
        for page_idx, page in enumerate(doc):
            page_width = page.rect.width
            raw_words = page.get_text("words")
            
            words = []
            for w in raw_words:
                words.append({
                    "text": w[4],
                    "x": w[0],
                    "y": w[1],
                    "norm_x": w[0] / page_width,
                    "page": page_idx
                })
            page_lines = group_words_into_lines(words, tolerance=6)
            lines.extend(page_lines)
        doc.close()
        return lines
    else:
        doc.close()
        pages = convert_from_path(pdf_path, dpi=300, poppler_path=POPPLER_PATH)
        for page_idx, page in enumerate(pages):
            page_width, _ = page.size
            data = pytesseract.image_to_data(page, output_type=Output.DICT, lang="eng")
            
            words = []
            for i in range(len(data["text"])):
                txt = data["text"][i].strip()
                if txt == "":
                    continue
                words.append({
                    "text": txt,
                    "x": data["left"][i],
                    "y": data["top"][i],
                    "norm_x": data["left"][i] / page_width,
                    "page": page_idx
                })
            page_lines = group_words_into_lines(words, tolerance=15)
            lines.extend(page_lines)
        return lines

# =====================================================
# COLUMNS DETECTION (RELATIVE SCALE)
# =====================================================

def detect_columns(lines):
    qty_x, unit_x, amount_x = None, None, None
    
    for line in lines:
        txt = " ".join(w["text"] for w in line).upper()
        if "CATEGORY" in txt and "QTY" in txt:
            for w in line:
                t = w["text"].upper().strip(":.,-")
                if qty_x is None and "QTY" in t:
                    qty_x = w["norm_x"]
                elif unit_x is None and "UNIT" in t:
                    unit_x = w["norm_x"]
                elif amount_x is None and "AMOUNT" in t:
                    amount_x = w["norm_x"]
            break
            
    if qty_x is None: qty_x = 0.48
    if unit_x is None: unit_x = 0.58
    if amount_x is None: amount_x = 0.70
    
    return {"qty": qty_x, "unit": unit_x, "amount": amount_x}

# =====================================================
# TEXT PARSERS (DATE, PLANT, WORK TYPE)
# =====================================================

def pdf_text(pdf):
    text = ""
    try:
        doc = fitz.open(pdf)
        for page in doc:
            text += page.get_text()
        doc.close()
    except Exception:
        text = ""
    
    if not text.strip():
        pages = convert_from_path(pdf, dpi=300, poppler_path=POPPLER_PATH)
        for img in pages:
            text += pytesseract.image_to_string(img, lang="eng") + "\n"
            
    return clean(text)

def extract_date(text):
    m = re.search(r"Date\s*:?\s*\|?\s*(\d{1,2}[\/\.-]\d{1,2}[\/\.-]\d{4})", text, re.I)
    if m:
        return m.group(1).replace(".", "/").replace("-", "/")
    m = re.search(r"\b(\d{1,2}[\/\.-]\d{1,2}[\/\.-]\d{4})\b", text)
    if m:
        return m.group(1).replace(".", "/").replace("-", "/")
    return ""

def extract_plant_no(text):
    m = re.search(r"Plant\s+Number\s*:?\s*\|?\s*([A-Z0-9]+)", text, re.I)
    if m:
        return m.group(1).strip()
    m = re.search(r"\b(YC\d{3}|YAB\d{3}|SAB\d{4}|YSL\d{3,4})\b", text, re.I)
    if m:
        return m.group(1).strip()
    return ""

def extract_work_type(text):
    checked_patterns = [
        (r"(?:☑|☒|\[x\]|\[X\])\s*Monthly\s+Servicing", "Maintenance"),
        (r"(?:☑|☒|\[x\]|\[X\])\s*Breakdown\s+Repair", "Repair"),
        (r"(?:☑|☒|\[x\]|\[X\])\s*Workshop\s+Repair", "Repair"),
        (r"(?:☑|☒|\[x\]|\[X\])\s*Monthly\s+Checking", "Maintenance"),
        (r"(?:☑|☒|\[x\]|\[X\])\s*Return\s+to\s+Workshop", "Repair"),
    ]
    for pattern, work_type in checked_patterns:
        if re.search(pattern, text, re.I):
            return work_type
            
    t = text.upper()
    if "MONTHLY SERVICING" in t or "MAINTENANCE" in t:
        return "Maintenance"
    if "BREAKDOWN REPAIR" in t or "WORKSHOP REPAIR" in t or "REPAIR" in t:
        return "Repair"
    return ""

# =====================================================
# COMPLETE INVOICE PARSER WITH SECTION TRACKING
# =====================================================

def parse_invoice(pdf):
    text = pdf_text(pdf)
    lines = get_pdf_words_and_lines(pdf)
    cols = detect_columns(lines)
    
    active_section = None
    rows = []
    
    for line in lines:
        if not line:
            continue
            
        line_text = " ".join(w["text"] for w in line).upper()
        
        # 1. Stop parsing if we hit the end of the tables
        if len(line_text.strip()) >= 5 and ("TOTAL" in line_text or "REMARK" in line_text):
            break
            
        # 2. Section Header Detection - Always evaluated first
        is_header = False
        if "SPARE PARTS" in line_text or "EXTERNAL SERVICES" in line_text:
            active_section = "spare"
            is_header = True
        elif "MATERIALS" in line_text:
            active_section = "material"
            is_header = True
        elif "LABOURS" in line_text or "LABOUR" in line_text or "LABORS" in line_text or "LABOR" in line_text:
            active_section = "labour"
            is_header = True
        elif "TRANSPORTATION" in line_text or "TRANSPORT" in line_text:
            active_section = "transport"
            is_header = True
            
        if is_header:
            continue  # Move directly to next line; do not process header as data
            
        # 3. For a valid data row, it must start with a line number index (1, 2, 3...)
        first_word = line[0]["text"]
        if not re.match(r"^\d+$", first_word.strip(".")):
            continue
            
        # 4. Extract potential financial values for the current row
        qty_val, unit_val, amount_val = None, None, None
        category_words = []
        
        for w in line:
            nx = w["norm_x"]
            txt = w["text"]
            
            if txt in ("$", "S$", "☐", "☑", "☒", "\\$"):
                continue
                
            if nx < cols["qty"] - 0.05:
                category_words.append(txt)
            elif nx < cols["unit"] - 0.04:
                if is_number(txt): qty_val = money(txt)
            elif nx < cols["amount"] - 0.04:
                if is_number(txt): unit_val = money(txt)
            else:
                if is_number(txt): amount_val = money(txt)
                
        if qty_val is None and unit_val is None and amount_val is None:
            continue
            
        if active_section is None:
            active_section = "spare"  # Failsafe default
            
        category_str = " ".join(category_words).strip()
        category_str = re.sub(r"^\d+\.?\s*", "", category_str) # clean leading index
        
        if category_str.upper() in ("", "PC", "ITRS", "JOB", "TRIP"):
            continue
            
        rows.append({
            "section": active_section,
            "category": category_str,
            "qty": qty_val,
            "unit": unit_val,
            "amount": amount_val
        })
        
    # Build Itemized Multi-line Strings
    extracted = {
        "spare": {"qty": [], "cost": []},
        "material": {"qty": [], "cost": []},
        "labour": {"qty": [], "cost": []},
        "transport": {"qty": [], "cost": []}
    }
    
    for r in rows:
        sec = r["section"]
        qty = r["qty"] or 0
        unit = r["unit"] or 0
        amount = r["amount"] or 0
        
        if amount > 0:
            cost = amount
        elif qty > 0 and unit > 0:
            cost = qty * unit
        else:
            cost = 0.0
            
        if qty > 0 or cost > 0:
            fmt_qty = str(int(qty)) if float(qty).is_integer() else str(qty)
            fmt_cost = f"{cost:.2f}"
            extracted[sec]["qty"].append(fmt_qty)
            extracted[sec]["cost"].append(fmt_cost)
            
    results = {}
    for sec in extracted:
        results[f"{sec}_qty"] = "\n".join(extracted[sec]["qty"])
        results[f"{sec}_cost"] = "\n".join(extracted[sec]["cost"])
            
    return {
        "date": extract_date(text),
        "work": extract_work_type(text),
        "text": text,
        **results
    }

# =====================================================
# ALL PDF FINDER
# =====================================================

def all_pdfs():
    pdfs = []
    for root, _, files in os.walk(INVOICE_ROOT):
        for file in files:
            if file.lower().endswith(".pdf"):
                pdfs.append(os.path.join(root, file))
    pdfs.sort()
    return pdfs

# =====================================================
# PROCESS ENGINE
# =====================================================

wb = load_workbook(EXCEL_FILE)
ws = wb["July_Job_Log"]

file_col = None
for c in range(1, ws.max_column + 1):
    if str(ws.cell(8, c).value).strip().upper() == "FILE":
        file_col = c
        break

if file_col is None:
    raise Exception("File column not found on row 8.")

def next_row():
    row = 9  # Starts on row 9
    while True:
        plant = ws.cell(row, 3).value
        if plant in (None, ""):
            return row
        row += 1

def write_record(pdf):
    plant = os.path.basename(os.path.dirname(pdf)).upper()
    data = parse_invoice(pdf)
    
    asset = asset_type(plant)
    if not asset:
        text_plant = extract_plant_no(data["text"]).upper()
        if asset_type(text_plant):
            plant = text_plant
            asset = asset_type(plant)

    row = next_row()
    print("Writing to row:", row)
    sn = row - 8

    # Basic Info
    ws.cell(row, 1).value = sn
    ws.cell(row, 2).value = data["date"]
    ws.cell(row, 3).value = plant
    ws.cell(row, 4).value = asset
    ws.cell(row, 5).value = data["work"]

    # Wrap-text alignment to stack items vertically inside the cell
    wrap_align = Alignment(wrap_text=True, vertical="top")

    # Column Mapping Details
    cols_map = {
        6: "spare_qty", 7: "spare_cost",
        8: "material_qty", 9: "material_cost",
        10: "labour_qty", 11: "labour_cost",
        12: "transport_qty", 13: "transport_cost"
    }

    # Write Data Blocks Item by Item
    for col_idx, data_key in cols_map.items():
        cell = ws.cell(row, col_idx)
        cell.value = data[data_key]
        cell.alignment = wrap_align

    # Link writing logic
    filename = os.path.splitext(os.path.basename(pdf))[0]
    m = re.search(r"\d{2}\.\d{2}\.\d{4}", filename)
    display = m.group(0).replace(".", "/") if m else filename.replace(".", "/")

    cell = ws.cell(row, file_col)
    cell.value = display
    cell.hyperlink = pdf
    cell.font = Font(color="0563C1", underline="single")

    # Terminal output cleaner (replacing newlines with commas for readability in terminal)
    print(
        f"{sn} | {plant} | {data['date']} | {data['work']} | "
        f"SP=[{data['spare_qty'].replace(chr(10), ', ')}]/$[{data['spare_cost'].replace(chr(10), ', ')}] | "
        f"MT=[{data['material_qty'].replace(chr(10), ', ')}]/$[{data['material_cost'].replace(chr(10), ', ')}] | "
        f"LB=[{data['labour_qty'].replace(chr(10), ', ')}]/$[{data['labour_cost'].replace(chr(10), ', ')}] | "
        f"TR=[{data['transport_qty'].replace(chr(10), ', ')}]/$[{data['transport_cost'].replace(chr(10), ', ')}]"
    )

# Run entire sequence
pdfs = all_pdfs()
print("=" * 60)
print("TOTAL PDFS FOUND:", len(pdfs))
print("=" * 60)

for pdf in pdfs:
    try:
        print("\n-----------------------------------------")
        print(os.path.basename(pdf))
        write_record(pdf)
    except Exception as e:
        print("\nFAILED:", pdf)
        print(e)

print("=" * 60)
print("Saving workbook...")
wb.save(EXCEL_FILE)
wb.close()
print("Done.")
