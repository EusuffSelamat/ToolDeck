from openpyxl import load_workbook
from openpyxl.styles import Font
from playwright.sync_api import sync_playwright, TimeoutError
from urllib.parse import urlparse, parse_qs
import pyperclip

EXCEL_FILE = r"C:\Users\eusuff.binselamat\Downloads\Updated Full List Equipment.xlsm"

wb = load_workbook(EXCEL_FILE, keep_vba=True)

SPIDER_URL = "https://onedrive.live.com/my?id=%2Fpersonal%2F29e3ab39e591aab9%2FDocuments%2FService%20Report%2FSpider%20Crane&viewid=505e8630%2D2f3c%2D459b%2Db55e%2Dae65e86cee0d"

BOOM_URL = "https://onedrive.live.com/my?id=%2Fpersonal%2F29e3ab39e591aab9%2FDocuments%2FService%20Report%2FBoom%20Lift&viewid=505e8630%2D2f3c%2D459b%2Db55e%2Dae65e86cee0d"

SCISSOR_URL = "https://onedrive.live.com/my?id=%2Fpersonal%2F29e3ab39e591aab9%2FDocuments%2FService%20Report%2FScissor%20Lift&viewid=505e8630%2D2f3c%2D459b%2Db55e%2Dae65e86cee0d"


# -----------------------------
# Excel
# -----------------------------

ws = wb.active

headers = {}

for c in ws[1]:
    if c.value:
        headers[str(c.value).strip()] = c.column

plant_col = headers["Plant No."]
if "Hyperlink" in headers:
    hyperlink_col = headers["Hyperlink"]
else:
    hyperlink_col = ws.max_column + 1
    ws.cell(1, hyperlink_col).value = "Hyperlink"
    wb.save(EXCEL_FILE)
ServiceReport_col = headers["Service Report"]

# -----------------------------
# Helper Functions
# -----------------------------

def get_page(plant):

    if plant.startswith(("YAB", "SAB")):
        return boom, BOOM_URL

    if plant.startswith("YSL"):
        return scissor, SCISSOR_URL

    return spider, SPIDER_URL


def goto_home(page, home):

    page.goto(home)

    page.wait_for_url(home + "*", timeout=15000)

    page.wait_for_load_state("domcontentloaded")

    page.wait_for_timeout(3000)


def search_folder(page, plant):

    search = page.get_by_role("searchbox", name="Search")

    search.click()

    search.press("Control+A")

    search.press("Backspace")

    search.fill(plant)

    search.press("Enter")

    page.wait_for_load_state("networkidle")

    page.wait_for_timeout(2500)


def open_folder(page, plant, home):

    folder = page.get_by_role("link", name=plant)

    folder.first.wait_for(timeout=10000)

    old = page.url

    folder.first.click()

    page.wait_for_function(
        "(url)=>window.location.href!=url",
        arg=old,
        timeout=10000
    )

    page.wait_for_load_state("networkidle")
    page.wait_for_timeout(1000)

    # Build the full folder URL including viewid
    current = page.url

    parsed = urlparse(current)
    params = parse_qs(parsed.query)

    folder_id = params["id"][0]

    home_parsed = urlparse(home)
    home_params = parse_qs(home_parsed.query)

    viewid = home_params["viewid"][0]

def copy_share_link(page):

    # Clear clipboard first
    pyperclip.copy("")

    # Click Copy link
    page.get_by_role("menuitem", name="Copy link").click()

    # Wait until clipboard receives the link
    link = ""

    for _ in range(30):      # up to 15 seconds

        page.wait_for_timeout(500)

        link = pyperclip.paste().strip()

        if link.startswith("http"):
            break

    if not link.startswith("http"):
        raise Exception("Failed to copy share link.")

    # Close the Share dialog
    page.locator("iframe[name='shareFrame']").content_frame\
        .get_by_role("button", name="Close").click()

    page.wait_for_timeout(500)

    return link

def save_link(row, link):

    # Hidden URL column
    url_cell = ws.cell(row=row, column=hyperlink_col)
    url_cell.value = link

    # Service Report column
    service_cell = ws.cell(row=row, column=ServiceReport_col)

    service_cell.value = "📂 Open Folder"

    # Dummy hyperlink (only to get the hand cursor)
    service_cell.hyperlink = "#"

    service_cell.font = Font(
        color="0563C1",
        underline="single"
    )

    wb.save(EXCEL_FILE)


# -----------------------------
# Browser
# -----------------------------

with sync_playwright() as p:

    browser = p.chromium.launch(
        channel="msedge",
        headless=False
    )

    context = browser.new_context()

    spider = context.new_page()
    spider.goto(SPIDER_URL)
    input("Login Spider Crane then press ENTER...")

    boom = context.new_page()
    boom.goto(BOOM_URL)
    input("Login Boom Lift then press ENTER...")

    scissor = context.new_page()
    scissor.goto(SCISSOR_URL)
    input("Login Scissor Lift then press ENTER...")

    for excel_row in range(2, ws.max_row + 1):

        existing = ws.cell(excel_row, ServiceReport_col).value

        if existing not in (None, ""):
            print(f"Skipping row {excel_row}")
            continue

        plant = str(ws.cell(excel_row, plant_col).value or "").strip()
        plant = plant.split("(")[0].strip()

        if not plant:
            continue

        if existing not in (None, ""):
            print(f"Skipping {plant}")
            continue

        print("=" * 60)
        print(f"Processing {plant}")

        try:

            page, home = get_page(plant)

            page.bring_to_front()

            goto_home(page, home)

            success = False

            # Retry search up to 3 times
            for attempt in range(3):

                print(f"Search attempt {attempt+1}")

                search_folder(page, plant)

                folder = page.get_by_text(plant, exact=True)

                if folder.count() > 0:
                    success = True
                    break

                page.wait_for_timeout(3000)

            if not success:
                print(f"Folder not found: {plant}")
                continue

            try:

                open_folder(page, plant, home)

                link = copy_share_link(page)

            except TimeoutError:

                print("Folder opened slowly... waiting longer.")

                page.wait_for_timeout(5000)

                current = page.url

                parsed = urlparse(current)
                params = parse_qs(parsed.query)

                folder_id = params["id"][0]

                link = (
                    "https://onedrive.live.com/my?"
                    f"id={folder_id}"
                    "&viewid=505e8630-2f3c-459b-b55e-ae65e86cee0d"
                )

            print(link)

            save_link(excel_row, link)

            cell = ws.cell(excel_row, hyperlink_col)
            print("Cell value:", cell.value)
            print("Hyperlink:", cell.hyperlink.target if cell.hyperlink else "None")

            print(f"✓ Saved {plant}")

            # Return to root folder
            goto_home(page, home)

        except Exception as e:

            print(f"ERROR processing {plant}")

            print(e)

            try:
                goto_home(page, home)
            except:
                pass

            continue

    browser.close()

wb.close()

print()
print("=" * 60)
print("Finished!")
print("=" * 60)

print("Finished.")
